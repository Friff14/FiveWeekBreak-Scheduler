import mimetypes
import os

import falcon
import json

from data.tables import *
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension

DBSession = sessionmaker(bind=engine)

alignment = Alignment(horizontal='center',
                      vertical='top')
highlight = NamedStyle(name="highlight")
fill = PatternFill("solid", fgColor="B3A2C7")
highlight.fill = fill
side = Side(border_style="thin", color="000000")
border = Border(left=side, right=side, top=side, bottom=side)
highlight.border = border
highlight.alignment = alignment

light_highlight = NamedStyle(name="lightHighlight")
fill = PatternFill("solid", fgColor="D9D9D9")
light_highlight.fill = fill
light_highlight.border = border
light_highlight.alignment = alignment

cols = 'ABCDEFGHIJKLMN'


class xlsx_creation:

    def on_get(self, req, resp, semester=None):
        resp.status = falcon.HTTP_200
        resp.set_header('Content-type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp.location = '/exports/' + str(semester)
        self.make_xlsx_file(semester)
        name = 'output_' + str(semester) + '.xlsx'
        # image_path = os.path.join(self.storage_path, name)
        resp.stream = open(name, 'rb')
        resp.stream_len = os.path.getsize(name)

    def make_xlsx_file(self, semester):
        session = DBSession()

        wb = Workbook()

        wb.add_named_style(highlight)

        ws = wb.active

        ws.append(
            ['INSTRUCTOR', 'COURSE', '', 'CRN', 'HOURS', '', 'DAYS',
             'ROOM', 'MAX', 'HRS', 'CAMP', 'Pay', 'Load/Ovld', 'Hrs Req'])

        ws.merge_cells('B1:C1')
        ws.merge_cells('E1:F1')

        for col in cols:
            ws[col + '1'].style = highlight

        instructors = session.query(Instructor).filter(Section.semester_id == semester) \
            .order_by(Instructor.instructor_last_name)

        current_row = 2

        for instructor in instructors:
            first = True
            hrs_req = instructor.instructor_hours_required
            total_hours = 0
            instructor_row_count = 0
            for section in instructor.sections:
                hours = ['', '']
                days = ''
                hrs = 0
                for time in section.schedule_times:
                    hours = [time.schedule_time_start_time, time.schedule_time_end_time]
                    days += time.schedule_time_day_of_week[0]
                    hrs += math.ceil(time.calc_length())
                total_hours += hrs
                # print(section.room.building.campus.campus_name)
                loadovld = ''
                if total_hours <= hrs_req:
                    loadovld = 'FL'
                else:
                    ovld = total_hours - hrs_req
                    loadovld = 'FL-' + str(max(0, int(hrs - ovld))) + '/FO-' + str(int(ovld))
                ws.append([instructor.instructor_first_name + ' ' + instructor.instructor_last_name if first else '',
                           section.course.prefix.prefix_name,
                           section.course.course_number,
                           section.section_crn,
                           hours[0],
                           hours[1],
                           days,
                           section.room.room_name,
                           section.section_capacity,
                           hrs,
                           section.room.building.campus.campus_name,
                           ' ',  # PAY - what is this?
                           loadovld,
                           hrs_req if first else ''
                           ])

                first = False

                current_row += 1
                instructor_row_count += 1

            for release in instructor.release:
                if total_hours <= hrs_req:
                    loadovld = ''
                else:
                    ovld = total_hours - hrs_req
                    loadovld = ' --- ' + str(int(ovld)) + ' hours overload'

                ws.append(['',
                           release.release_name + ' --- ' +
                           str(int(release.release_hours)) + ' units release' +
                           loadovld
                           ])

                ws.merge_cells('B' + str(current_row) + ':M' + str(current_row))
                ws['B' + str(current_row)].style = highlight
                current_row += 1
                instructor_row_count += 1

            ws.merge_cells(start_column=1, end_column=1, start_row=current_row - instructor_row_count,
                           end_row=current_row - 1)
            ws['A' + str(current_row - instructor_row_count)].style = light_highlight

        filename = 'output_' + str(semester) + '.xlsx'
        wb.save(filename)

        return {"url": filename}
